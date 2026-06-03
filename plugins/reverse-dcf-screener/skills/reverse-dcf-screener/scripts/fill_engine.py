#!/usr/bin/env python3
"""Reverse DCF (Terminal-Anchored) engine — fill input cells + compute in parallel.
Reads JSON from stdin, prints JSON result to stdout. Keeps Excel formulas intact
so the file recalculates when opened in Excel."""
import sys, json, os, shutil, datetime, warnings
warnings.filterwarnings("ignore")

TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets", "reverse_dcf_screener.xlsx")

# field -> Engine input cell (yellow). DO NOT touch formula cells.
ENGINE_CELLS = {
    "ticker":"C4","revenue_r0":"C5","ev":"C6","sector":"C7","wacc_override":"C9",
    "terminal_margin":"C11","terminal_g":"C12","terminal_roic":"C13","tax":"C14",
    "horizon_n":"C15","hist_cagr":"C18","fade":"C19","tam":"C20","max_pen":"C21",
    "abs_ceiling":"C22","buffer":"C40","price":"C43","shares_m":"C44","net_debt":"C45",
    "analyst_target":"C48","analyst_range":"C49","consensus_fy1":"C50",
}

# --- Screener sheet (master comparison table) ---
SCREENER_FIRST_ROW = 10          # data rows start at 10 (row 9 = header)
SCREENER_FORMULA_ROW = 10        # source row to translate formulas from when growing past template
# field -> Screener input column (letter). Formula cols C/L/M/N/O/P are NEVER written.
SCREENER_INPUT_COLS = {
    "ticker":"A","sector":"B","revenue_r0":"D","ev":"E","terminal_margin":"F",
    "terminal_g":"G","terminal_roic":"H","horizon_n":"I","hist_cagr":"J","tam":"K",
}
SCREENER_FORMULA_COLS = ("C","L","M","N","O","P")  # Excel recalcs these — leave intact
# global assumptions shared by ALL screener rows (so Gaps are comparable)
GLOBAL_CELLS = {"tax":"S3","fade":"S4","max_pen":"S5","abs_ceiling":"S6","buffer":"S7"}
GLOBAL_DEFAULTS = {"tax":0.21,"fade":0.70,"max_pen":0.25,"abs_ceiling":0.45,"buffer":0.05}

DEP_HELP = (
    "ERROR: ต้องมี openpyxl ก่อนถึงจะเขียนไฟล์ Excel ได้\n"
    "  pip install openpyxl\n"
    "ถ้าเจอ 'externally-managed-environment' (PEP 668 — macOS homebrew / Linux):\n"
    "  python3 -m venv .venv && . .venv/bin/activate && pip install openpyxl\n"
    "  # หรือ:  pipx run --spec openpyxl ...   หรือ:  pip install --break-system-packages openpyxl\n"
    "หมายเหตุ: compute()/verdict/screener-view ไม่ต้องใช้ openpyxl — จำเป็นเฉพาะตอนเขียนไฟล์\n"
)


def _require_openpyxl():
    """Lazy import with a clear, actionable message instead of a raw traceback."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        sys.stderr.write(DEP_HELP)
        sys.exit(1)

def compute(d):
    R0=d["revenue_r0"]; EV=d["ev"]; W=d["wacc"]; m=d["terminal_margin"]
    g=d["terminal_g"]; roic=d["terminal_roic"]; tax=d["tax"]; N=d["horizon_n"]
    hist=d["hist_cagr"]; fade=d["fade"]; tam=d.get("tam",0); maxpen=d["max_pen"]
    absc=d["abs_ceiling"]; buf=d["buffer"]; nd=d["net_debt"]; sh=d["shares_m"]
    fwd = (d["consensus_fy1"]/R0 - 1) if d.get("consensus_fy1") else 0
    reinv = g/roic
    tv = EV*(1+W)**N
    fcff = tv*(W-g)
    conv = m*(1-tax)*(1-reinv)
    rstar = fcff/conv
    implied = (rstar/R0)**(1/(N+1)) - 1
    capA = max(hist, fwd)*fade
    capB = 999 if tam in (0,None) else (maxpen*tam/R0)**(1/(N+1)) - 1
    capC = absc
    plausible = min(capA, capB, capC)
    gap = implied - plausible
    if gap > buf: verdict = "แพง — Priced for Perfection"
    elif gap < -buf: verdict = "ถูก — Low Expectations"
    else: verdict = "Fair — สมเหตุสมผล"
    def price_at(cagr):
        return ((R0*(1+cagr)**(N+1)*conv/(W-g)/(1+W)**N) - nd)*1000/sh
    zones = {
        "strong_buy": price_at(max(plausible-0.05,0)),
        "fair_value": price_at(plausible),
        "caution_low": price_at(plausible+0.05),
        "caution_high": price_at(plausible+0.10),
        "red_flag": price_at(plausible+0.10),
    }
    return {
        "ticker": d.get("ticker"), "reinvestment": reinv, "terminal_value": tv,
        "fcff_n1": fcff, "conversion": conv, "implied_terminal_revenue": rstar,
        "implied_cagr": implied, "cap_a": capA, "cap_b": capB, "cap_c": capC,
        "plausible_cagr": plausible, "gap": gap, "verdict": verdict,
        "ev_sales": EV/R0, "current_price": d.get("price"), "zones": zones,
    }

def write_xlsx(d, out_path):
    openpyxl = _require_openpyxl()
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)
    eng = wb["Engine"]
    eng["C7"] = d.get("sector")
    if "wacc" in d:
        eng["C9"] = d["wacc"]   # put WACC as override (guard against placeholder sector lookup)
    for f, cell in ENGINE_CELLS.items():
        if f in ("sector","wacc_override"): continue
        if f in d and d[f] is not None:
            eng[cell] = d[f]
    wb.save(out_path)
    return out_path


def _ensure_screener_formulas(ws, row):
    """Rows past the template's pre-filled block have no formulas. Translate them
    from SCREENER_FORMULA_ROW so a freshly-appended row still recalcs in Excel.
    Only fills EMPTY formula cells — never overwrites an existing formula."""
    from openpyxl.formula.translate import Translator
    for col in SCREENER_FORMULA_COLS:
        cell = ws[f"{col}{row}"]
        if cell.value not in (None, ""):
            continue
        src = ws[f"{col}{SCREENER_FORMULA_ROW}"].value
        if isinstance(src, str) and src.startswith("="):
            origin = f"{col}{SCREENER_FORMULA_ROW}"
            cell.value = Translator(src, origin=origin).translate_formula(f"{col}{row}")


def append_screener(d, screener_path):
    """Append one stock as a new row in the master Screener sheet. Writes input
    cols only (formula cols recalc in Excel). Returns (path, row, effective_globals).
    Globals (S3-S7) are written once if blank, then reused so all rows are comparable."""
    openpyxl = _require_openpyxl()
    if not os.path.exists(screener_path):
        shutil.copy(TEMPLATE, screener_path)
    wb = openpyxl.load_workbook(screener_path)
    ws = wb["Screener"]

    # resolve + persist shared global assumptions (existing value wins, else payload, else default)
    globals_eff = {}
    for field, cell in GLOBAL_CELLS.items():
        cur = ws[cell].value
        if cur in (None, ""):
            val = d.get(field, GLOBAL_DEFAULTS[field])
            ws[cell] = val
            globals_eff[field] = val
        else:
            globals_eff[field] = cur

    # find next free data row (Ticker col A is input — empty => available)
    row = SCREENER_FIRST_ROW
    while ws[f"A{row}"].value not in (None, ""):
        row += 1
    _ensure_screener_formulas(ws, row)

    # write input cols only
    for field, col in SCREENER_INPUT_COLS.items():
        val = d.get(field)
        if val is not None:
            ws[f"{col}{row}"] = val
    wacc_override = d.get("wacc_override", d.get("wacc"))
    if wacc_override is not None:
        ws[f"Q{row}"] = wacc_override

    wb.save(screener_path)
    return screener_path, row, globals_eff

def main():
    d = json.load(sys.stdin)
    if d.get("mode") == "screener" or d.get("screener_file"):
        screener_path = d.get("screener_file") or os.path.join("analyses", "screener.xlsx")
        parent = os.path.dirname(screener_path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        path, row, globals_eff = append_screener(d, screener_path)
        # recompute with the file's effective globals so the chat table matches Excel recalc
        res = compute({**d, **globals_eff})
        res["screener_file"] = path
        res["screener_row"] = row
        res["screener_globals"] = globals_eff
    else:
        res = compute(d)
        if not d.get("no_write"):
            os.makedirs("analyses", exist_ok=True)
            stamp = d.get("date") or datetime.date.today().isoformat()
            out = os.path.join("analyses", f"{d.get('ticker','STOCK')}_{stamp}.xlsx")
            res["file"] = write_xlsx(d, out)
    print(json.dumps(res, ensure_ascii=False))

if __name__ == "__main__":
    main()

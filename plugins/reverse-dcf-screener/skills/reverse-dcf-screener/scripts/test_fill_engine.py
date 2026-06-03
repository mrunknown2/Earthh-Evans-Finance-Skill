import json, subprocess, sys, os, tempfile, shutil
HERE = os.path.dirname(os.path.abspath(__file__))
SCRIPT = os.path.join(HERE, "fill_engine.py")

IREN = {
    "ticker": "IREN", "revenue_r0": 0.757, "ev": 22.6, "sector": "Software (System/Application)",
    "wacc": 0.095, "terminal_margin": 0.35, "terminal_g": 0.04, "terminal_roic": 0.15,
    "tax": 0.25, "horizon_n": 10, "hist_cagr": 0.60, "fade": 0.70, "tam": 150.0,
    "max_pen": 0.25, "abs_ceiling": 0.45, "buffer": 0.05, "price": 65.33,
    "shares_m": 357.38, "net_debt": -0.7, "consensus_fy1": 1.09,
    "analyst_target": 79.04, "analyst_range": "$29 - $100+", "no_write": True
}

def run(payload):
    p = subprocess.run([sys.executable, SCRIPT], input=json.dumps(payload),
                       capture_output=True, text=True)
    assert p.returncode == 0, p.stderr
    return json.loads(p.stdout)

def approx(a, b, tol=0.005): return abs(a-b) <= tol

def test_iren_golden():
    r = run(IREN)
    assert approx(r["implied_cagr"], 0.320), r["implied_cagr"]
    assert approx(r["plausible_cagr"], 0.420), r["plausible_cagr"]
    assert approx(r["gap"], -0.100), r["gap"]
    assert "ถูก" in r["verdict"], r["verdict"]
    assert approx(r["zones"]["fair_value"], 143.56, tol=0.6), r["zones"]["fair_value"]
    assert approx(r["zones"]["strong_buy"], 97.42, tol=0.6), r["zones"]["strong_buy"]
    assert approx(r["zones"]["red_flag"], 301.30, tol=1.0), r["zones"]["red_flag"]

def test_tam_zero_uses_other_caps():
    payload = dict(IREN); payload["tam"] = 0
    r = run(payload)
    assert approx(r["plausible_cagr"], 0.420), r["plausible_cagr"]

def test_expensive_verdict():
    payload = dict(IREN); payload["price"] = 250.0; payload["ev"] = 90.0
    r = run(payload)
    assert "แพง" in r["verdict"], (r["verdict"], r["implied_cagr"])


# --- Screener mode (append to master Screener sheet) ---

# a second, distinct stock to verify multi-row append (no_write removed implicitly)
SECOND = {
    "ticker": "TSTB", "revenue_r0": 2.0, "ev": 30.0, "sector": "Retail (General)",
    "wacc": 0.09, "terminal_margin": 0.12, "terminal_g": 0.03, "terminal_roic": 0.20,
    "tax": 0.25, "horizon_n": 10, "hist_cagr": 0.15, "fade": 0.70, "tam": 80.0,
    "max_pen": 0.25, "abs_ceiling": 0.45, "buffer": 0.05, "net_debt": 1.0,
    "shares_m": 100.0, "consensus_fy1": 2.2,
}


def _screener_ws(path):
    import openpyxl
    return openpyxl.load_workbook(path)["Screener"]


def test_screener_appends_two_rows():
    d = tempfile.mkdtemp()
    try:
        sf = os.path.join(d, "screener.xlsx")
        r1 = run({**IREN, "screener_file": sf})
        r2 = run({**SECOND, "screener_file": sf})
        assert r1["screener_row"] == 10, r1["screener_row"]
        assert r2["screener_row"] == 11, r2["screener_row"]
        ws = _screener_ws(sf)
        assert ws["A10"].value == "IREN", ws["A10"].value
        assert ws["A11"].value == "TSTB", ws["A11"].value
        # input cols written on row 10 (IREN)
        assert approx(ws["D10"].value, 0.757), ws["D10"].value
        assert approx(ws["E10"].value, 22.6), ws["E10"].value
        assert approx(ws["F10"].value, 0.35), ws["F10"].value
        assert ws["I10"].value == 10, ws["I10"].value
        assert approx(ws["K10"].value, 150.0), ws["K10"].value
        assert approx(ws["Q10"].value, 0.095), ws["Q10"].value  # WACC override
        assert ws["B10"].value == "Software (System/Application)", ws["B10"].value
    finally:
        shutil.rmtree(d)


def test_screener_chat_drops_forward_cagr():
    # The Screener sheet has NO consensus/forward column — its Cap A formula is
    # hist*fade only. So in screener mode the chat result must drop forward CAGR
    # (compute with fwd=0) to match what Excel recalcs. Here forward (1.8/1.0-1=0.80)
    # is far above hist (0.10); with forward kept, plausible would be ~0.34, but the
    # Screener sheet would show hist*fade = 0.10*0.70 = 0.07.
    FWD_HEAVY = {
        "ticker": "FWDX", "revenue_r0": 1.0, "ev": 20.0, "sector": "Software",
        "wacc": 0.09, "terminal_margin": 0.30, "terminal_g": 0.04, "terminal_roic": 0.15,
        "tax": 0.25, "horizon_n": 10, "hist_cagr": 0.10, "fade": 0.70, "tam": 100.0,
        "max_pen": 0.25, "abs_ceiling": 0.45, "buffer": 0.05, "net_debt": 0.0,
        "shares_m": 100.0, "consensus_fy1": 1.8,  # forward = 0.80 >> hist 0.10
    }
    d = tempfile.mkdtemp()
    try:
        sf = os.path.join(d, "screener.xlsx")
        r = run({**FWD_HEAVY, "screener_file": sf})
        # file globals: fade=0.70 -> Cap A = hist*fade = 0.07 (forward must be dropped)
        assert approx(r["plausible_cagr"], 0.07, tol=0.01), r["plausible_cagr"]
        assert approx(r["cap_a"], 0.07, tol=0.01), r["cap_a"]
    finally:
        shutil.rmtree(d)


def test_screener_keeps_formula_cols_intact():
    d = tempfile.mkdtemp()
    try:
        sf = os.path.join(d, "screener.xlsx")
        run({**IREN, "screener_file": sf})
        ws = _screener_ws(sf)
        for col in ("C", "L", "M", "N", "O", "P"):
            v = ws[f"{col}10"].value
            assert isinstance(v, str) and v.startswith("="), (col, v)
    finally:
        shutil.rmtree(d)


def test_screener_creates_file_from_template():
    d = tempfile.mkdtemp()
    try:
        sf = os.path.join(d, "nested", "screener.xlsx")
        r = run({**IREN, "screener_file": sf})
        assert os.path.exists(sf), "screener file not created"
        assert r["screener_file"] == sf, r["screener_file"]
        import openpyxl
        wb = openpyxl.load_workbook(sf)
        assert "Screener" in wb.sheetnames, wb.sheetnames
    finally:
        shutil.rmtree(d)


def test_write_requires_openpyxl_gracefully():
    # simulate openpyxl missing: a None entry in sys.modules makes `import` raise
    # ImportError. compute()/no_write must still work; only write needs the dep.
    code = (
        "import sys; sys.modules['openpyxl'] = None; "
        f"sys.path.insert(0, {HERE!r}); "
        "import fill_engine; "
        "fill_engine.write_xlsx({'sector': 'x', 'wacc': 0.1}, '/tmp/_should_not_exist.xlsx')"
    )
    p = subprocess.run([sys.executable, "-c", code], capture_output=True, text=True)
    assert p.returncode == 1, (p.returncode, p.stderr)
    assert "openpyxl" in p.stderr, p.stderr
    assert "pip install" in p.stderr, p.stderr
    assert "Traceback" not in p.stderr, p.stderr  # clean message, not a raw stack


def test_compute_works_without_openpyxl():
    # the analytical path (verdict / screener-view) must stay openpyxl-free
    code = (
        "import sys, json; sys.modules['openpyxl'] = None; "
        f"sys.path.insert(0, {HERE!r}); "
        "import fill_engine; "
        "r = fill_engine.compute(json.loads(sys.argv[1])); "
        "print(r['verdict'])"
    )
    p = subprocess.run([sys.executable, "-c", code, json.dumps(IREN)],
                       capture_output=True, text=True)
    assert p.returncode == 0, p.stderr
    assert "ถูก" in p.stdout, p.stdout


def test_screener_uses_file_globals_for_chat():
    # template global S3 tax = 0.21; IREN payload tax = 0.25.
    # effective tax must be the file's existing global (0.21) so the chat table
    # matches what Excel will recalc from S3.
    d = tempfile.mkdtemp()
    try:
        sf = os.path.join(d, "screener.xlsx")
        r = run({**IREN, "screener_file": sf})
        assert approx(r["screener_globals"]["tax"], 0.21), r["screener_globals"]
        ws = _screener_ws(sf)
        assert approx(ws["S3"].value, 0.21), ws["S3"].value  # untouched
    finally:
        shutil.rmtree(d)


if __name__ == "__main__":
    import traceback
    tests = [v for k, v in sorted(globals().items())
             if k.startswith("test_") and callable(v)]
    passed = failed = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
            passed += 1
        except Exception as e:
            print(f"FAIL  {t.__name__}: {e}")
            traceback.print_exc()
            failed += 1
    print(f"\n{passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
